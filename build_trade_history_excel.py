"""
build_trade_history_excel.py — Converts data/trade_history.json to Excel.
Run locally: python3 build_trade_history_excel.py
Creates: data/trade_history.xlsx with 2 tabs:
  1. By Date    — all trades sorted newest first
  2. By Insider — all trades sorted by ticker/insider then date
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = "data/trade_history.json"
OUTPUT_FILE = "data/trade_history.xlsx"

HEADERS = [
    "Date", "Ticker", "Insider Name", "Title",
    "Shares", "Price", "Total Value ($)", "Shares Owned After",
    "Source"
]

COL_WIDTHS = [12, 8, 28, 30, 12, 10, 16, 18, 10]

HEADER_FILL  = PatternFill("solid", start_color="0F1117", end_color="0F1117")
HEADER_FONT  = Font(name="Arial", bold=True, color="00BA7C", size=10)
ROW_FONT     = Font(name="Arial", size=9)
ALT_FILL     = PatternFill("solid", start_color="0A0A0A", end_color="0A0A0A")
BORDER_SIDE  = Side(style="thin", color="2F2F2F")
CELL_BORDER  = Border(bottom=Side(style="thin", color="2F2F2F"))


def load_trades():
    with open(INPUT_FILE) as f:
        history = json.load(f)

    rows = []
    for key, trades in history.items():
        parts  = key.split(":", 1)
        ticker = parts[0] if len(parts) > 0 else ""
        insider = parts[1] if len(parts) > 1 else ""

        if not isinstance(trades, list):
            continue
        for t in trades:
            if not t.get("is_buy") and t.get("code", "P") != "P":
                continue
            rows.append({
                "date":        t.get("date", ""),
                "ticker":      ticker,
                "insider":     t.get("insider_name", insider),
                "title":       t.get("title", ""),
                "shares":      t.get("shares", 0),
                "price":       t.get("price_per_share", 0),
                "total":       t.get("total_value", 0),
                "owned":       t.get("shares_owned", 0),
                "source":      t.get("source", "live"),
            })
    return rows


def write_sheet(ws, rows, title):
    ws.title = title

    # Header row
    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Data rows
    for i, row in enumerate(rows, 2):
        fill = ALT_FILL if i % 2 == 0 else None
        values = [
            row["date"],
            row["ticker"],
            row["insider"],
            row["title"],
            row["shares"],
            row["price"],
            row["total"],
            row["owned"] if row["owned"] else "",
            row["source"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = Font(name="Arial", size=9, color="E7E9EA")
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            # Format numbers
            if col == 5:  # Shares
                cell.number_format = "#,##0"
            elif col == 6:  # Price
                cell.number_format = "$#,##0.00"
            elif col in (7, 8):  # Total, Owned
                cell.number_format = "#,##0"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # Tab color
    ws.sheet_properties.tabColor = "00BA7C"


def main():
    print(f"Loading {INPUT_FILE}...")
    rows = load_trades()
    print(f"  {len(rows):,} trade records loaded")

    # Tab 1: By Date (newest first)
    by_date = sorted(rows, key=lambda x: x["date"], reverse=True)

    # Tab 2: By Insider (ticker > insider > date)
    by_insider = sorted(rows, key=lambda x: (x["ticker"], x["insider"], x["date"]))

    wb = Workbook()
    ws1 = wb.active
    write_sheet(ws1, by_date, "By Date")

    ws2 = wb.create_sheet()
    write_sheet(ws2, by_insider, "By Insider")

    # Summary sheet
    ws3 = wb.create_sheet("Summary")
    ws3.sheet_properties.tabColor = "1D9BF0"
    ws3["A1"] = "Form4Wire Trade History"
    ws3["A1"].font = Font(name="Arial", bold=True, size=14, color="1D9BF0")
    ws3["A3"] = "Generated:"
    ws3["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws3["A4"] = "Total Records:"
    ws3["B4"] = len(rows)
    ws3["A5"] = "Unique Tickers:"
    ws3["B5"] = len(set(r["ticker"] for r in rows))
    ws3["A6"] = "Unique Insiders:"
    ws3["B6"] = len(set(f"{r['ticker']}:{r['insider']}" for r in rows))
    ws3["A7"] = "Date Range:"
    ws3["B7"] = f"{min(r['date'] for r in rows if r['date'])} → {max(r['date'] for r in rows if r['date'])}"
    for row in range(3, 8):
        ws3[f"A{row}"].font = Font(name="Arial", bold=True, size=10, color="71767B")
        ws3[f"B{row}"].font = Font(name="Arial", size=10, color="E7E9EA")
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 35
    ws3.sheet_view.showGridLines = False

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"✅ Saved to {OUTPUT_FILE}")
    print(f"   Tab 1 'By Date':     {len(by_date):,} rows")
    print(f"   Tab 2 'By Insider':  {len(by_insider):,} rows")


if __name__ == "__main__":
    main()
